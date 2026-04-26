import os
import json
import asyncio
import threading
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from engine import FormSender

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs('uploads', exist_ok=True)
os.makedirs('outputs', exist_ok=True)

# 進捗を保持するグローバル変数
progress_store = {}

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'ファイルが選択されていません'}), 400
    path = os.path.join(app.config['UPLOAD_FOLDER'], 'companies.xlsx')
    file.save(path)

    # プレビュー用にデータ読み込み
    df = pd.read_excel(path, sheet_name='協力会社リスト', header=8)
    df.columns = ['No', '会社名', 'URL']
    df = df.dropna(subset=['会社名', 'URL'])
    companies = df[['会社名', 'URL']].to_dict(orient='records')

    # 送信者情報読み込み
    wb = load_workbook(path)
    ws = wb['協力会社リスト']
    sender = {
        '送信者名': ws['B2'].value,
        '送信者会社名': ws['B3'].value,
        'メール': ws['B4'].value,
        '電話': ws['B5'].value,
        '本文': ws['B6'].value,
    }
    return jsonify({'companies': companies, 'sender': sender, 'count': len(companies)})

@app.route('/run', methods=['POST'])
def run():
    data = request.json
    mode = data.get('mode', 'test')  # test or production
    session_id = datetime.now().strftime('%Y%m%d%H%M%S')
    progress_store[session_id] = {'status': 'running', 'results': [], 'total': 0, 'done': 0}

    def run_async():
        asyncio.run(execute(session_id, mode))

    thread = threading.Thread(target=run_async)
    thread.start()
    return jsonify({'session_id': session_id})

@app.route('/progress/<session_id>')
def progress(session_id):
    return jsonify(progress_store.get(session_id, {'status': 'not_found'}))

@app.route('/download/<session_id>')
def download(session_id):
    path = f'outputs/result_{session_id}.xlsx'
    if os.path.exists(path):
        return send_file(path, as_attachment=True, download_name='送信結果.xlsx')
    return jsonify({'error': 'ファイルが見つかりません'}), 404

async def execute(session_id, mode):
    try:
        wb = load_workbook('uploads/companies.xlsx')
        ws = wb['協力会社リスト']
        sender = {
            'name': ws['B2'].value or '',
            'company': ws['B3'].value or '',
            'email': ws['B4'].value or '',
            'phone': ws['B5'].value or '',
            'message': ws['B6'].value or '',
        }
        document_url = ws['B7'].value or ''  # B7に資料URLを追加

        df = pd.read_excel('uploads/companies.xlsx', sheet_name='協力会社リスト', header=8)
        df.columns = ['No', '会社名', 'URL']
        df = df.dropna(subset=['会社名', 'URL'])
        companies = df.to_dict(orient='records')

        progress_store[session_id]['total'] = len(companies)
        results = []

        sender_obj = FormSender(test_mode=(mode == 'test'))
        await sender_obj.init()

        for i, company in enumerate(companies):
            result = await sender_obj.process(company['会社名'], company['URL'], sender, document_url)
            results.append(result)
            progress_store[session_id]['done'] = i + 1
            progress_store[session_id]['results'] = results

        await sender_obj.close()
        save_result(session_id, results)
        progress_store[session_id]['status'] = 'done'

    except Exception as e:
        progress_store[session_id]['status'] = 'error'
        progress_store[session_id]['error'] = str(e)

def save_result(session_id, results):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    thin = Side(style='thin', color='AAAAAA')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 全件結果シート
    ws1 = wb.active
    ws1.title = '全件結果'
    headers = ['No.', '会社名', 'ステータス', '失敗理由', '手動対応要否', '送信フォームURL', '実行日時']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='1F4E79')
        cell.border = border

    status_colors = {'✅ 送信完了': 'E2EFDA', '❌ 失敗': 'FFE0E0', '⚠️ スキップ': 'FFF2CC'}
    for i, r in enumerate(results, 1):
        row = [i, r['company'], r['status'], r.get('reason', '-'),
               '要手動対応' if r['status'] != '✅ 送信完了' else '-',
               r.get('form_url', '-'), r.get('timestamp', '-')]
        color = status_colors.get(r['status'], 'FFFFFF')
        for col, val in enumerate(row, 1):
            cell = ws1.cell(row=i+1, column=col, value=val)
            cell.fill = PatternFill('solid', start_color=color)
            cell.border = border

    # 手動対応リストシート
    ws2 = wb.create_sheet('手動対応リスト')
    h2 = ['No.', '会社名', 'トップページURL', '失敗理由', '対応済み？']
    for col, h in enumerate(h2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color='843C0C')
        cell.border = border

    manual = [r for r in results if r['status'] != '✅ 送信完了']
    for i, r in enumerate(manual, 1):
        row = [i, r['company'], r.get('top_url', '-'), r.get('reason', '-'), '□']
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=i+1, column=col, value=val)
            cell.fill = PatternFill('solid', start_color='FFF2CC')
            cell.border = border

    for ws in [ws1, ws2]:
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 16
        if ws == ws1:
            ws.column_dimensions['F'].width = 40
            ws.column_dimensions['G'].width = 20

    wb.save(f'outputs/result_{session_id}.xlsx')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
